"""
===============================================================================
FICOM_VIZ.PY - Structure GHT & Visualisation
===============================================================================
DIM - Data Intelligence Médicale
Nettoyage des fichiers Excel de structure GHT et génération d'arbres visuels
===============================================================================
"""

import json
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass, field
from enum import Enum

import polars as pl
from loguru import logger

# Imports optionnels pour la visualisation
try:
    import networkx as nx
    NETWORKX_AVAILABLE = True
except ImportError:
    NETWORKX_AVAILABLE = False
    logger.warning("NetworkX non disponible - pip install networkx")

try:
    from graphviz import Digraph
    GRAPHVIZ_AVAILABLE = True
except ImportError:
    GRAPHVIZ_AVAILABLE = False
    logger.warning("Graphviz non disponible - pip install graphviz")

try:
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    logger.warning("Matplotlib non disponible - pip install matplotlib")


# =============================================================================
# CONFIGURATION
# =============================================================================

class NodeType(Enum):
    """Types de noeuds dans la structure GHT."""
    GHT = "GHT"                      # Groupement Hospitalier de Territoire
    ETABLISSEMENT = "ETABLISSEMENT"   # Établissement membre
    POLE = "POLE"                     # Pôle d'activité
    SERVICE = "SERVICE"               # Service/Unité fonctionnelle
    UM = "UM"                         # Unité médicale
    SITE = "SITE"                     # Site géographique


@dataclass
class GHTNode:
    """Représente un noeud dans la structure GHT."""
    id: str
    name: str
    node_type: NodeType
    parent_id: Optional[str] = None
    finess: Optional[str] = None
    attributes: Dict[str, Any] = field(default_factory=dict)
    children: List["GHTNode"] = field(default_factory=list)


@dataclass
class FICOMConfig:
    """Configuration du module FICOM."""
    output_dir: Path = field(default_factory=lambda: Path("./output"))
    graph_format: str = "png"  # png, svg, pdf
    graph_dpi: int = 150
    color_scheme: Dict[NodeType, str] = field(default_factory=lambda: {
        NodeType.GHT: "#1a5276",
        NodeType.ETABLISSEMENT: "#2874a6",
        NodeType.POLE: "#5dade2",
        NodeType.SERVICE: "#85c1e9",
        NodeType.UM: "#aed6f1",
        NodeType.SITE: "#d4e6f1"
    })


# =============================================================================
# PARSER EXCEL STRUCTURE GHT
# =============================================================================

class GHTStructureParser:
    """
    Parser pour les fichiers Excel de structure GHT.
    Gère les formats souvent "illisibles" des fichiers administratifs.
    """

    def __init__(self, config: FICOMConfig = None):
        self.config = config or FICOMConfig()
        self.nodes: List[GHTNode] = []
        self.hierarchy: Dict[str, GHTNode] = {}

    def parse_excel(self, filepath: Path) -> List[GHTNode]:
        """
        Parse un fichier Excel de structure GHT.
        Gère plusieurs formats possibles.
        """
        logger.info(f"Parsing structure GHT: {filepath}")

        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl requis - pip install openpyxl")
            return []

        # Lecture avec Polars (via openpyxl)
        try:
            df = pl.read_excel(filepath, engine="openpyxl")
        except Exception as e:
            logger.warning(f"Erreur lecture standard, tentative alternative: {e}")
            df = self._read_excel_fallback(filepath)

        if df is None or df.is_empty():
            logger.error(f"Impossible de lire {filepath}")
            return []

        logger.info(f"  Colonnes détectées: {df.columns}")
        logger.info(f"  {len(df)} lignes")

        # Nettoyage des données
        df = self._clean_dataframe(df)

        # Détection du format et extraction de la structure
        nodes = self._extract_hierarchy(df)

        self.nodes = nodes
        return nodes

    def _read_excel_fallback(self, filepath: Path) -> Optional[pl.DataFrame]:
        """
        Lecture alternative pour les fichiers Excel problématiques.
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb.active

            # Extraction des données brutes
            data = []
            headers = None

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                # Ignorer les lignes vides
                if all(cell is None for cell in row):
                    continue

                # Première ligne non vide = headers
                if headers is None:
                    headers = [str(cell) if cell else f"col_{i}" for i, cell in enumerate(row)]
                    continue

                # Données
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = cell
                data.append(row_dict)

            wb.close()

            if not data:
                return None

            return pl.DataFrame(data)

        except Exception as e:
            logger.error(f"Échec lecture fallback: {e}")
            return None

    def _clean_dataframe(self, df: pl.DataFrame) -> pl.DataFrame:
        """
        Nettoie le DataFrame des données Excel.
        """
        # Suppression des colonnes entièrement vides
        non_empty_cols = [
            col for col in df.columns
            if df.select(pl.col(col).is_not_null().any()).item()
        ]
        df = df.select(non_empty_cols)

        # Normalisation des noms de colonnes
        rename_map = {}
        for col in df.columns:
            clean_name = str(col).strip().upper()
            clean_name = clean_name.replace(" ", "_")
            clean_name = clean_name.replace("-", "_")
            rename_map[col] = clean_name

        df = df.rename(rename_map)

        # Suppression des lignes entièrement vides
        df = df.filter(
            pl.any_horizontal(pl.all().is_not_null())
        )

        return df

    def _extract_hierarchy(self, df: pl.DataFrame) -> List[GHTNode]:
        """
        Extrait la hiérarchie depuis le DataFrame nettoyé.
        """
        nodes = []
        columns = [c.upper() for c in df.columns]

        # Détection des colonnes clés
        col_mapping = {
            "ght": self._find_column(columns, ["GHT", "GROUPEMENT", "TERRITOIRE"]),
            "etablissement": self._find_column(columns, ["ETABLISSEMENT", "ETAB", "HOPITAL", "CH"]),
            "finess": self._find_column(columns, ["FINESS", "FINESS_PMSI", "CODE_FINESS"]),
            "pole": self._find_column(columns, ["POLE", "POLES"]),
            "service": self._find_column(columns, ["SERVICE", "UNITE", "UF"]),
            "um": self._find_column(columns, ["UM", "UNITE_MEDICALE", "CODE_UM"]),
            "site": self._find_column(columns, ["SITE", "SITE_GEO", "LOCALISATION"])
        }

        logger.info(f"  Mapping colonnes: {col_mapping}")

        # Construction de la hiérarchie
        ght_nodes = {}
        etab_nodes = {}
        pole_nodes = {}

        for row in df.iter_rows(named=True):
            # Niveau GHT
            ght_name = row.get(col_mapping["ght"]) if col_mapping["ght"] else None
            if ght_name and str(ght_name).strip():
                ght_id = f"GHT_{self._normalize_id(ght_name)}"
                if ght_id not in ght_nodes:
                    ght_node = GHTNode(
                        id=ght_id,
                        name=str(ght_name).strip(),
                        node_type=NodeType.GHT
                    )
                    ght_nodes[ght_id] = ght_node
                    nodes.append(ght_node)

            # Niveau Établissement
            etab_name = row.get(col_mapping["etablissement"]) if col_mapping["etablissement"] else None
            finess = row.get(col_mapping["finess"]) if col_mapping["finess"] else None

            if etab_name and str(etab_name).strip():
                etab_id = f"ETAB_{self._normalize_id(etab_name)}"
                if etab_id not in etab_nodes:
                    parent = list(ght_nodes.keys())[-1] if ght_nodes else None
                    etab_node = GHTNode(
                        id=etab_id,
                        name=str(etab_name).strip(),
                        node_type=NodeType.ETABLISSEMENT,
                        parent_id=parent,
                        finess=str(finess).strip() if finess else None
                    )
                    etab_nodes[etab_id] = etab_node
                    nodes.append(etab_node)

                    # Lier au parent GHT
                    if parent and parent in ght_nodes:
                        ght_nodes[parent].children.append(etab_node)

            # Niveau Pôle
            pole_name = row.get(col_mapping["pole"]) if col_mapping["pole"] else None
            if pole_name and str(pole_name).strip():
                pole_id = f"POLE_{self._normalize_id(pole_name)}"
                if pole_id not in pole_nodes:
                    parent = list(etab_nodes.keys())[-1] if etab_nodes else None
                    pole_node = GHTNode(
                        id=pole_id,
                        name=str(pole_name).strip(),
                        node_type=NodeType.POLE,
                        parent_id=parent
                    )
                    pole_nodes[pole_id] = pole_node
                    nodes.append(pole_node)

                    # Lier au parent Établissement
                    if parent and parent in etab_nodes:
                        etab_nodes[parent].children.append(pole_node)

            # Niveau Service/UM
            service_name = row.get(col_mapping["service"]) if col_mapping["service"] else None
            um_code = row.get(col_mapping["um"]) if col_mapping["um"] else None

            if service_name and str(service_name).strip():
                service_id = f"SVC_{self._normalize_id(service_name)}"
                parent = list(pole_nodes.keys())[-1] if pole_nodes else (
                    list(etab_nodes.keys())[-1] if etab_nodes else None
                )
                service_node = GHTNode(
                    id=service_id,
                    name=str(service_name).strip(),
                    node_type=NodeType.SERVICE,
                    parent_id=parent,
                    attributes={"um": str(um_code) if um_code else None}
                )
                nodes.append(service_node)

        self.hierarchy = {n.id: n for n in nodes}
        return nodes

    def _find_column(self, columns: List[str], keywords: List[str]) -> Optional[str]:
        """Trouve une colonne correspondant aux mots-clés."""
        for col in columns:
            for kw in keywords:
                if kw in col:
                    return col
        return None

    def _normalize_id(self, name: str) -> str:
        """Normalise un nom pour en faire un ID."""
        return str(name).strip().replace(" ", "_").replace("-", "_")[:30]

    def export_to_pmsi_pilot(self, output_path: Path = None) -> Path:
        """
        Exporte la structure dans un format injectable pour PMSI-Pilot.
        """
        if output_path is None:
            output_path = self.config.output_dir / "structure_ght_pmsi_pilot.json"

        self.config.output_dir.mkdir(parents=True, exist_ok=True)

        # Format PMSI-Pilot
        structure = {
            "version": "1.0",
            "generated_at": str(Path(__file__).stat().st_mtime),
            "hierarchy": []
        }

        for node in self.nodes:
            node_data = {
                "id": node.id,
                "name": node.name,
                "type": node.node_type.value,
                "parent_id": node.parent_id,
                "finess": node.finess,
                "attributes": node.attributes
            }
            structure["hierarchy"].append(node_data)

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(structure, f, indent=2, ensure_ascii=False)

        logger.info(f"  ✓ Export PMSI-Pilot: {output_path}")
        return output_path


# =============================================================================
# GÉNÉRATEUR DE GRAPHE VISUEL
# =============================================================================

class GHTGraphGenerator:
    """
    Génère des visualisations graphiques de la structure GHT.
    """

    def __init__(self, config: FICOMConfig = None):
        self.config = config or FICOMConfig()

    def generate_graphviz(
        self,
        nodes: List[GHTNode],
        output_name: str = "structure_ght"
    ) -> Optional[Path]:
        """
        Génère un graphe avec Graphviz.
        """
        if not GRAPHVIZ_AVAILABLE:
            logger.error("Graphviz non disponible")
            return None

        self.config.output_dir.mkdir(parents=True, exist_ok=True)

        # Création du graphe
        dot = Digraph(
            name='GHT Structure',
            format=self.config.graph_format,
            engine='dot'
        )

        # Configuration du style
        dot.attr(
            rankdir='TB',
            splines='ortho',
            nodesep='0.5',
            ranksep='0.8'
        )
        dot.attr('node', shape='box', style='rounded,filled', fontname='Arial')
        dot.attr('edge', color='#666666')

        # Ajout des noeuds
        for node in nodes:
            color = self.config.color_scheme.get(node.node_type, "#ffffff")
            label = f"{node.name}"
            if node.finess:
                label += f"\n({node.finess})"

            dot.node(
                node.id,
                label,
                fillcolor=color,
                fontcolor='white' if node.node_type in [NodeType.GHT, NodeType.ETABLISSEMENT] else 'black'
            )

        # Ajout des liens
        for node in nodes:
            if node.parent_id:
                dot.edge(node.parent_id, node.id)

        # Rendu
        output_path = self.config.output_dir / output_name
        try:
            dot.render(str(output_path), cleanup=True)
            full_path = Path(f"{output_path}.{self.config.graph_format}")
            logger.info(f"  ✓ Graphe Graphviz: {full_path}")
            return full_path
        except Exception as e:
            logger.error(f"Erreur rendu Graphviz: {e}")
            logger.info("  Assurez-vous que Graphviz est installé: https://graphviz.org/download/")
            return None

    def generate_networkx(
        self,
        nodes: List[GHTNode],
        output_name: str = "structure_ght_nx"
    ) -> Optional[Path]:
        """
        Génère un graphe avec NetworkX + Matplotlib.
        """
        if not NETWORKX_AVAILABLE or not MATPLOTLIB_AVAILABLE:
            logger.error("NetworkX ou Matplotlib non disponible")
            return None

        self.config.output_dir.mkdir(parents=True, exist_ok=True)

        # Création du graphe
        G = nx.DiGraph()

        # Ajout des noeuds avec attributs
        for node in nodes:
            G.add_node(
                node.id,
                label=node.name,
                node_type=node.node_type.value,
                color=self.config.color_scheme.get(node.node_type, "#cccccc")
            )

        # Ajout des arêtes
        for node in nodes:
            if node.parent_id:
                G.add_edge(node.parent_id, node.id)

        # Layout hiérarchique
        try:
            pos = nx.drawing.nx_agraph.graphviz_layout(G, prog='dot')
        except Exception:
            # Fallback si graphviz pas dispo pour nx
            pos = nx.spring_layout(G, k=2, iterations=50)

        # Dessin
        plt.figure(figsize=(16, 12), dpi=self.config.graph_dpi)

        # Couleurs des noeuds
        colors = [G.nodes[n].get('color', '#cccccc') for n in G.nodes()]
        labels = {n: G.nodes[n].get('label', n) for n in G.nodes()}

        nx.draw(
            G, pos,
            with_labels=True,
            labels=labels,
            node_color=colors,
            node_size=3000,
            font_size=8,
            font_weight='bold',
            arrows=True,
            edge_color='#666666',
            arrowsize=20
        )

        plt.title("Structure GHT", fontsize=14, fontweight='bold')
        plt.tight_layout()

        # Sauvegarde
        output_path = self.config.output_dir / f"{output_name}.{self.config.graph_format}"
        plt.savefig(output_path, dpi=self.config.graph_dpi, bbox_inches='tight')
        plt.close()

        logger.info(f"  ✓ Graphe NetworkX: {output_path}")
        return output_path

    def generate_html_tree(
        self,
        nodes: List[GHTNode],
        output_name: str = "structure_ght"
    ) -> Path:
        """
        Génère un arbre HTML interactif (fallback sans dépendances graphiques).
        """
        self.config.output_dir.mkdir(parents=True, exist_ok=True)

        # Construction de l'arbre HTML
        html_content = """<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <title>Structure GHT</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }
        h1 { color: #1a5276; }
        .tree { margin-left: 20px; }
        .node { margin: 5px 0; padding: 8px 12px; border-radius: 5px; display: inline-block; }
        .GHT { background: #1a5276; color: white; font-weight: bold; }
        .ETABLISSEMENT { background: #2874a6; color: white; }
        .POLE { background: #5dade2; color: white; }
        .SERVICE { background: #85c1e9; }
        .UM { background: #aed6f1; }
        .SITE { background: #d4e6f1; }
        .children { margin-left: 30px; border-left: 2px solid #ccc; padding-left: 15px; }
        .finess { font-size: 0.8em; opacity: 0.8; }
    </style>
</head>
<body>
    <h1>Structure GHT</h1>
    <div class="tree">
"""

        # Construction récursive de l'arbre
        root_nodes = [n for n in nodes if n.parent_id is None]

        def render_node(node: GHTNode, level: int = 0) -> str:
            html = f'<div class="node {node.node_type.value}">{node.name}'
            if node.finess:
                html += f' <span class="finess">({node.finess})</span>'
            html += '</div>\n'

            # Enfants
            children = [n for n in nodes if n.parent_id == node.id]
            if children:
                html += '<div class="children">\n'
                for child in children:
                    html += render_node(child, level + 1)
                html += '</div>\n'

            return html

        for root in root_nodes:
            html_content += render_node(root)

        html_content += """
    </div>
</body>
</html>
"""

        output_path = self.config.output_dir / f"{output_name}.html"
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        logger.info(f"  ✓ Arbre HTML: {output_path}")
        return output_path


# =============================================================================
# POINT D'ENTRÉE
# =============================================================================

def process_ght_structure(
    excel_path: Path,
    output_dir: Path = None,
    generate_graph: bool = True
) -> Tuple[List[GHTNode], Dict[str, Path]]:
    """
    Fonction principale pour traiter un fichier Excel de structure GHT.
    """
    output_dir = output_dir or Path("./output")
    config = FICOMConfig(output_dir=output_dir)

    # Parsing
    parser = GHTStructureParser(config)
    nodes = parser.parse_excel(excel_path)

    if not nodes:
        logger.warning("Aucune structure extraite")
        return [], {}

    logger.info(f"  {len(nodes)} noeuds extraits")

    outputs = {}

    # Export PMSI-Pilot
    outputs["pmsi_pilot"] = parser.export_to_pmsi_pilot()

    # Génération des graphes
    if generate_graph:
        generator = GHTGraphGenerator(config)

        # Graphviz
        gv_path = generator.generate_graphviz(nodes)
        if gv_path:
            outputs["graphviz"] = gv_path

        # NetworkX (backup)
        nx_path = generator.generate_networkx(nodes)
        if nx_path:
            outputs["networkx"] = nx_path

        # HTML (toujours généré comme fallback)
        html_path = generator.generate_html_tree(nodes)
        outputs["html"] = html_path

    return nodes, outputs


def main():
    """Point d'entrée principal."""
    import argparse

    parser = argparse.ArgumentParser(
        description="DIM FICOM Viz - Structure GHT et visualisation"
    )
    parser.add_argument(
        "input_file",
        type=str,
        help="Fichier Excel de structure GHT"
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default="./output",
        help="Répertoire de sortie"
    )
    parser.add_argument(
        "--format", "-f",
        type=str,
        choices=["png", "svg", "pdf"],
        default="png",
        help="Format du graphe"
    )
    parser.add_argument(
        "--no-graph",
        action="store_true",
        help="Ne pas générer de graphe visuel"
    )

    args = parser.parse_args()

    nodes, outputs = process_ght_structure(
        Path(args.input_file),
        Path(args.output),
        generate_graph=not args.no_graph
    )

    print(f"\n✓ Traitement terminé")
    print(f"  {len(nodes)} noeuds dans la structure")
    for name, path in outputs.items():
        print(f"  {name}: {path}")


if __name__ == "__main__":
    main()
